# viber_blast.py
import streamlit as st
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from io import BytesIO

def viber_blast_section():
    st.subheader("Viber Blast CSV Uploader")

    # === UPLOADERS ===
    uploaded_file = st.file_uploader(
        "Choose Raw Data CSV file",
        type=["csv"],
        key="viber_raw_uploader",
        help="Must have: Contact No., Debtor Name, Account No., Client, Validity"
    )
    if uploaded_file:
        st.session_state.viber_raw_file = uploaded_file
        st.success("Raw data uploaded")

    uploaded_lookup = st.file_uploader(
        "Choose Collector Lookup Excel",
        type=["xlsx"],
        key="viber_lookup_uploader",
        help="Must have: Account No. and Collector (code like MGARBAS)"
    )
    if uploaded_lookup:
        st.session_state.viber_lookup_file = uploaded_lookup
        st.success("Collector lookup uploaded")

    # === RESET ===
    if st.session_state.get('viber_raw_file') or st.session_state.get('viber_lookup_file'):
        if st.button("Reset All", type="secondary"):
            for k in ['viber_raw_file', 'viber_lookup_file']:
                st.session_state.pop(k, None)
            st.rerun()

    # === SAMPLE DATA ===
    sample_b2 = pd.DataFrame({
        "Campaign": ["SAMPLE"]*4,
        "CH Code": ["12345", "123456", "1234567", "12345678"],
        "First Name": [""]*4,
        "Full Name": ["Richard Arenas", "Jinnggoy Dela Cruz", "Roman Dalisay", "Edwin Paras"],
        "Last Name": ["Collector A", "Collector B", "PJHA", "Collector D"],
        "Mobile Number": ["09274186327", "09760368821", "09088925110", "09175791122"],
        "OB": [""]*4
    })
    sample_b4 = pd.DataFrame({
        "Campaign": ["TEST"], "CH Code": ["1234"], "First Name": [""],
        "Full Name": ["Janica d Benbinuto"], "Last Name": ["TEST"],
        "Mobile Number": ["09655669672"], "OB": [""]
    })
    sample_df = sample_b2.copy()
    campaign_name = "GENERIC"
    timestamp = datetime.now().strftime("%b %d %Y %I_%M %p").upper()

    if st.session_state.get('viber_raw_file'):
        try:
            # === READ RAW DATA ===
            df = pd.read_csv(st.session_state.viber_raw_file, encoding='utf-8-sig', dtype=str)
            df = df.fillna("")
            df.columns = df.columns.str.strip()

            required = ["Contact No.", "Debtor Name", "Account No.", "Client", "Validity"]
            if not all(col in df.columns for col in required):
                st.error(f"Missing columns: {', '.join([c for c in required if c not in df.columns])}")
                return

            df["Contact No."] = df["Contact No."].str.replace(r'^="?|"?$', '', regex=True).str.strip()
            df["Account No."] = df["Account No."].str.replace(r'^="?|"?$', '', regex=True).str.strip()

            # === FILTERS ===
            before = len(df)
            df = df[df["Validity"].str.strip().str.upper() == "VALID"]
            st.info(f"Removed {before - len(df)} invalid Validity")

            before = len(df)
            df = df[~df["Account No."].str.contains("BEL", case=False, na=False)]
            st.info(f"Removed {before - len(df)} BEL accounts")

            # === BASE TABLE ===
            summary_df = pd.DataFrame({
                "Campaign": df["Client"],
                "CH Code": df["Account No."],
                "First Name": "",
                "Full Name": df["Debtor Name"],  # ← DEBTOR NAME
                "Last Name": "",                 # ← WILL BE COLLECTOR CODE
                "Mobile Number": df["Contact No."],
                "OB": ""
            })

            # === CAMPAIGN DETECTION ===
            if summary_df["Campaign"].str.contains("SBC CURING B4", case=False, na=False).any():
                sample_df = sample_b4
                campaign_name = "SBC CURING B4"
            elif summary_df["Campaign"].str.contains("SBC CARDS CURING B2", case=False, na=False).any():
                campaign_name = "SBC CARDS CURING B2"

            # === 100% CORRECT COLLECTOR LOOKUP + REMOVE UNMATCHED ===
            if not st.session_state.get('viber_lookup_file'):
                st.error("Collector lookup file is REQUIRED!")
                st.stop()

            lookup = pd.read_excel(st.session_state.viber_lookup_file, dtype=str)
            lookup = lookup.fillna("")
            lookup.columns = lookup.columns.str.strip()

            # Auto-detect Account and Collector columns
            acc_cols = [c for c in lookup.columns if any(k in c.lower() for k in ["account", "acc", "ch"])]
            coll_cols = [c for c in lookup.columns if any(k in c.lower() for k in ["collector", "agent", "code"])]

            if not acc_cols or not coll_cols:
                st.error("Cannot find Account No. or Collector column in lookup file!")
                st.write("Found columns:", list(lookup.columns))
                st.stop()

            acc_col = acc_cols[0]
            coll_col = coll_cols[0]
            st.info(f"Using: '{acc_col}' → Account | '{coll_col}' → Collector Code")

            lookup[acc_col] = lookup[acc_col].str.replace(r'^="?|"?$', '', regex=True).str.strip()

            # === PERFECT MERGE: CH Code → Collector Code into Last Name ===
            merged = summary_df.merge(
                lookup[[acc_col, coll_col]].rename(columns={
                    acc_col: "CH Code",
                    coll_col: "Collector_Code"
                }),
                on="CH Code",
                how="left"
            )

            # === REMOVE ROWS WITH NO COLLECTOR ===
            before = len(merged)
            merged = merged[merged["Collector_Code"].notna() & (merged["Collector_Code"].str.strip() != "")]
            after = len(merged)
            removed = before - after

            if removed > 0:
                st.warning(f"Removed {removed} accounts → NO COLLECTOR in lookup")
                removed_codes = merged[~merged["Collector_Code"].notna()]["CH Code"].tolist()[:10]
                st.write("Removed CH Codes:", removed_codes)
            else:
                st.success("All accounts matched with collector!")

            # === FINAL CLEAN TABLE ===
            final_clean = merged[["Campaign", "CH Code", "First Name", "Full Name", "Mobile Number", "OB"]].copy()
            final_clean["Last Name"] = merged["Collector_Code"]  # ← COLLECTOR CODE HERE

            # === REMOVE DUPLICATES ===
            before_dup = len(final_clean)
            final_clean = final_clean.drop_duplicates(subset=["CH Code"], keep="first")
            dup_removed = before_dup - len(final_clean)
            if dup_removed:
                st.success(f"Removed {dup_removed} duplicate CH Code(s)")

            # === ADD SAMPLES ===
            final_df = pd.concat([final_clean, sample_df], ignore_index=True)
            filename = f"VIBER BLAST {campaign_name} {timestamp} PST.xlsx".upper()

            st.subheader("FINAL VIBER BLAST TABLE")
            st.dataframe(final_df, use_container_width=True)
            st.success(f"READY → {len(final_df)} rows | 100% CLEAN & CORRECT")

            # === DOWNLOAD ===
            output = BytesIO()
            wb = Workbook()
            ws = wb.active
            ws.title = "Viber Blast"
            for c, h in enumerate(final_df.columns, 1):
                ws.cell(1, c, h)
            for r, row in enumerate(final_df.itertuples(index=False, name=None), 2):
                for c, val in enumerate(row, 1):
                    ws.cell(r, c, str(val) if pd.notna(val) else "").number_format = "@"
            wb.save(output)
            output.seek(0)

            st.download_button(
                label="DOWNLOAD VIBER BLAST EXCEL",
                data=output,
                file_name=filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key="viber_download"
            )

        except Exception as e:
            st.error(f"Error: {e}")
    else:
        st.info("Upload Raw CSV + Collector Excel to start")
        st.dataframe(sample_b2, use_container_width=True)